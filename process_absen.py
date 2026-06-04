#!/usr/bin/env python3
"""
Attendance Data Processor
=========================
Reads an HTML attendance report exported from the KSPS attendance system,
extracts the table data, and fills an Excel template while preserving
all existing formatting (fonts, borders, fills, etc.).

Output columns:
  A = Date (1–31)
  B = Check-in time (HH:MM) — or holiday/weekend text (merged B:E, gray fill)
  C = Check-out time (HH:MM)
  D = (empty)
  E = (empty)

Usage:
    python3 process_absen.py \
        --html "Absen Bulan 03-2026.html" \
        --template template-table-absen.xlsx \
        --output output_absen.xlsx

Dependencies:
    pip install beautifulsoup4 openpyxl
"""

import argparse
import calendar
import copy
import logging
import os
import re
import sys
from datetime import date
from typing import Optional

from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Logging setup
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# Check for PIL / Pillow installation
try:
    import PIL
except ImportError:
    log.warning("=" * 80)
    log.warning("WARNING: Pillow (PIL) is not installed in the python environment.")
    log.warning("Without Pillow, openpyxl CANNOT parse and preserve images/logos from the Excel template!")
    log.warning("Please run 'pip install pillow' on your machine to restore template images.")
    log.warning("=" * 80)


# ---------------------------------------------------------------------------
# Data models
# ---------------------------------------------------------------------------
class AttendanceRecord:
    """Represents one day of attendance data parsed from the HTML table."""

    def __init__(
        self,
        tgl: int,
        masuk: Optional[str] = None,
        pulang: Optional[str] = None,
        keterangan: Optional[str] = None,
    ):
        self.tgl = tgl                    # Day of month (1-31)
        self.masuk = masuk                 # Check-in time (HH:MM) or None
        self.pulang = pulang               # Check-out time (HH:MM) or None
        self.keterangan = keterangan       # Holiday/leave note or None

    @property
    def is_holiday_or_leave(self) -> bool:
        """True when the row is a holiday / leave day with a text note."""
        return self.keterangan is not None

    @property
    def has_attendance(self) -> bool:
        """True when the employee has at least a check-in record."""
        return self.masuk is not None

    def __repr__(self) -> str:
        if self.is_holiday_or_leave:
            return f"Day {self.tgl:2d}: {self.keterangan}"
        if self.has_attendance:
            return f"Day {self.tgl:2d}: in={self.masuk} out={self.pulang}"
        return f"Day {self.tgl:2d}: (no data)"


# ---------------------------------------------------------------------------
# HTML Parsing
# ---------------------------------------------------------------------------
def _clean_text(text: Optional[str]) -> Optional[str]:
    """Strip whitespace and return None for empty / dash-only values."""
    if text is None:
        return None
    text = text.strip()
    if text in ("", "-"):
        return None
    return text


def _is_holiday_text(raw: Optional[str]) -> Optional[str]:
    """
    Detect holiday / leave text in the 'Terlambat' column.
    Returns the cleaned text if it's a non-numeric note, else None.
    """
    cleaned = _clean_text(raw)
    if cleaned is None:
        return None

    # If it matches the "N Menit" pattern or is purely numeric → not a note
    if re.match(r"^\s*\d+\s*(Menit)?\s*$", cleaned, re.IGNORECASE):
        return None

    # It's a textual note (e.g. "Cuti Bersama Hari Nyepi")
    return cleaned


def _extract_month_year(soup: BeautifulSoup) -> tuple[int, int]:
    """
    Extract the month and year from the HTML page title.
    Expected format in <title>: "Absen Bulan MM-YYYY"

    Returns
    -------
    tuple[int, int]
        (year, month)

    Raises
    ------
    ValueError
        If the month/year cannot be parsed from the title.
    """
    title_tag = soup.find("title")
    title_text = title_tag.get_text(strip=True) if title_tag else ""

    # Try to match "MM-YYYY" pattern in the title
    match = re.search(r"(\d{1,2})-(\d{4})", title_text)
    if match:
        month = int(match.group(1))
        year = int(match.group(2))
        log.info("Detected period from title: %04d-%02d", year, month)
        return year, month

    # Fallback: try <h1> tag (e.g. "Absen Bulan 03-2026")
    h1_tag = soup.find("h1")
    h1_text = h1_tag.get_text(strip=True) if h1_tag else ""
    match = re.search(r"(\d{1,2})-(\d{4})", h1_text)
    if match:
        month = int(match.group(1))
        year = int(match.group(2))
        log.info("Detected period from h1: %04d-%02d", year, month)
        return year, month

    raise ValueError(
        f"Cannot extract month/year from HTML. "
        f"Title: '{title_text}', H1: '{h1_text}'"
    )


def parse_html(html_path: str) -> tuple[list[AttendanceRecord], int, int]:
    """
    Parse the attendance HTML file and return attendance records with period.

    Parameters
    ----------
    html_path : str
        Path to the HTML file exported from the attendance system.

    Returns
    -------
    tuple[list[AttendanceRecord], int, int]
        (records, year, month) — one record per day found in the table.

    Raises
    ------
    FileNotFoundError
        If the HTML file does not exist.
    ValueError
        If the expected table structure is not found.
    """
    if not os.path.isfile(html_path):
        raise FileNotFoundError(f"HTML file not found: {html_path}")

    log.info("Reading HTML file: %s", html_path)
    with open(html_path, "r", encoding="utf-8") as fh:
        soup = BeautifulSoup(fh, "html.parser")

    # Extract month and year for weekend detection
    year, month = _extract_month_year(soup)

    # Locate the attendance table
    table = soup.find("table")
    if table is None:
        raise ValueError("No <table> element found in the HTML file.")

    tbody = table.find("tbody")
    if tbody is None:
        raise ValueError("No <tbody> element found inside the table.")

    rows = tbody.find_all("tr")
    log.info("Found %d rows in the attendance table.", len(rows))

    records: list[AttendanceRecord] = []

    for row in rows:
        cells = row.find_all("td")
        if len(cells) < 5:
            log.warning("Skipping row with fewer than 5 cells: %s",
                        row.get_text(strip=True)[:60])
            continue

        # Extract raw text from columns (ignore Aksi column at index 5)
        raw_tgl = _clean_text(cells[0].get_text())
        raw_masuk = _clean_text(cells[1].get_text())
        raw_pulang = _clean_text(cells[2].get_text())
        raw_terlambat = cells[3].get_text()  # keep raw for holiday detection

        # Parse date number
        if raw_tgl is None or not raw_tgl.isdigit():
            log.warning("Skipping row with invalid date: '%s'", raw_tgl)
            continue
        tgl = int(raw_tgl)

        # Detect holiday / leave text in the terlambat column
        holiday_text = _is_holiday_text(raw_terlambat)

        if holiday_text and raw_masuk is None and raw_pulang is None:
            # This is a holiday / leave row
            records.append(AttendanceRecord(tgl=tgl, keterangan=holiday_text))
            log.debug("Day %d → holiday/leave: %s", tgl, holiday_text)
        else:
            # Normal attendance row (may have partial data)
            records.append(AttendanceRecord(
                tgl=tgl,
                masuk=raw_masuk,
                pulang=_clean_text(cells[2].get_text()),
            ))
            log.debug("Day %d → masuk=%s pulang=%s", tgl, raw_masuk, raw_pulang)

    # Validate row count
    days_in_month = calendar.monthrange(year, month)[1]
    if len(records) != days_in_month:
        log.warning(
            "Expected %d date rows but found %d. "
            "Some dates may be missing in the HTML.",
            days_in_month, len(records),
        )

    return records, year, month


# ---------------------------------------------------------------------------
# Excel Writing
# ---------------------------------------------------------------------------

# Column mapping (defaults for fallback/legacy simple template)
COL_TGL = 1     # Column A — date number
COL_MASUK = 2   # Column B — check-in time
COL_PULANG = 3  # Column C — check-out time
COL_D = 4       # Column D — empty
COL_E = 5       # Column E — empty

# Thin border matching the template styling
_THIN_SIDE = Side(style="thin")
THIN_BORDER = Border(
    left=_THIN_SIDE,
    right=_THIN_SIDE,
    top=_THIN_SIDE,
    bottom=_THIN_SIDE,
)


def _unmerge_row(ws, row: int) -> None:
    """Unmerge any merged cell ranges that overlap the given row."""
    to_remove = []
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= row <= merged_range.max_row:
            to_remove.append(merged_range)

    for mr in to_remove:
        ws.unmerge_cells(str(mr))


def _merge_be(ws, row: int, start_col: int) -> None:
    """Merge cells from column start_col+1 to start_col+4 in the given row."""
    ws.merge_cells(
        start_row=row, start_column=start_col + 1,
        end_row=row, end_column=start_col + 4,
    )


def _copy_style_from_template(ws, template_ws, row: int, col: int) -> None:
    """
    Copy the cell style from the template to the working sheet.
    Useful when cells lose styling after unmerge operations.
    """
    src = template_ws.cell(row=row, column=col)
    dst = ws.cell(row=row, column=col)
    dst.font = copy.copy(src.font)
    dst.fill = copy.copy(src.fill)
    dst.border = copy.copy(src.border)
    dst.alignment = copy.copy(src.alignment)
    dst.number_format = src.number_format


def _apply_holiday_style(ws, row: int, start_col: int) -> None:
    """Apply gray fill + bold font + thin border to a holiday/weekend merged row section."""
    # Static gray fill supporting all office suites and HTML rendering
    gray_fill = PatternFill(
        start_color="E0E0E0",
        end_color="E0E0E0",
        fill_type="solid",
    )
    
    # Configure style for the primary text cell (start_col + 1)
    text_cell = ws.cell(row=row, column=start_col + 1)
    text_cell.font = Font(
        name="Palatino Linotype",
        size=10,
        bold=True,
        color=Color(theme=1),
        charset=134,
    )
    text_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )
    
    # Apply border and gray fill to all cells in the row section
    for col in range(start_col, start_col + 5):
        c = ws.cell(row=row, column=col)
        c.border = THIN_BORDER
        c.fill = gray_fill


def _apply_normal_style(ws, row: int, col: int) -> None:
    """Apply normal (non-holiday) styling + thin border to an attendance cell."""
    cell = ws.cell(row=row, column=col)
    cell.fill = PatternFill()  # No fill
    cell.font = Font(
        name="Palatino Linotype",
        size=10,
        bold=False,
        color=Color(indexed=8),
        charset=134,
    )
    cell.alignment = Alignment(horizontal="center")
    cell.border = THIN_BORDER


def _get_weekend_label(year: int, month: int, day: int) -> Optional[str]:
    """
    Return a weekend label if the given date falls on Saturday or Sunday.

    Returns
    -------
    str or None
        "LIBUR - SABTU" for Saturday, "LIBUR - MINGGU" for Sunday, else None.
    """
    try:
        d = date(year, month, day)
    except ValueError:
        return None  # Invalid date (e.g. Feb 30)

    weekday = d.weekday()  # Monday=0, Sunday=6
    if weekday == 5:
        return "LIBUR - SABTU"
    if weekday == 6:
        return "LIBUR - MINGGU"
    return None


def find_table_coordinates(ws) -> tuple[Optional[int], Optional[int]]:
    """
    Find the header row index and start column index of the attendance table dynamically.
    Looks for the cell containing 'Tgl' or 'Tanggal'. Once found, looks below it
    to find where the number 1 starts.
    """
    for r in range(1, 100):
        for c in range(1, 20):
            val = ws.cell(row=r, column=c).value
            if val and str(val).strip().lower() in ("tgl", "tanggal"):
                # Found 'Tgl' cell, now look below it to find the start of the table numbers
                for offset in range(1, 5):
                    test_val = ws.cell(row=r + offset, column=c).value
                    if test_val == 1 or str(test_val) == "1":
                        # The header row is the row immediately before the first day number
                        return r + offset - 1, c
                # Fallback: assume header is at r
                return r, c
    return None, None


INDONESIAN_MONTHS = {
    1: "JANUARI", 2: "FEBRUARI", 3: "MARET", 4: "APRIL", 
    5: "MEI", 6: "JUNI", 7: "JULI", 8: "AGUSTUS", 
    9: "SEPTEMBER", 10: "OKTOBER", 11: "NOPEMBER", 12: "DESEMBER"
}


def update_periode_cell(ws, year: int, month: int) -> None:
    """Find the cell containing 'PERIODE' and update the period string."""
    last_day = calendar.monthrange(year, month)[1]
    month_name = INDONESIAN_MONTHS.get(month, "")
    new_periode_val = f": 1 - {last_day} {month_name} {year}"
    
    for r in range(1, 50):
        for c in range(1, 15):
            val = ws.cell(row=r, column=c).value
            if val and str(val).strip().lower() == "periode":
                # Check cells to the right for the value (columns c+1 to c+6)
                for next_c in range(c + 1, c + 7):
                    right_val = ws.cell(row=r, column=next_c).value
                    if right_val is not None:
                        right_str = str(right_val).strip()
                        if right_str:
                            # Does it look like a period value cell?
                            looks_like_period = (
                                right_str.startswith(":") or 
                                right_str[0].isdigit() or
                                any(m in right_str.upper() for m in INDONESIAN_MONTHS.values()) or
                                "202" in right_str
                            )
                            if looks_like_period:
                                ws.cell(row=r, column=next_c).value = new_periode_val
                                log.info("Updated period in adjacent cell (%d, %d) to %s", r, next_c, new_periode_val)
                                return
                
                # Fallback: Update inside this cell itself if no adjacent value cell was found
                val_str = str(val)
                if ":" in val_str:
                    prefix = val_str.split(":")[0]
                    ws.cell(row=r, column=c).value = f"{prefix}{new_periode_val}"
                else:
                    ws.cell(row=r, column=c).value = f"PERIODE {new_periode_val}"
                log.info("Updated period in cell (%d, %d) to: %s", r, c, ws.cell(row=r, column=c).value)
                return


def write_to_excel(
    records: list[AttendanceRecord],
    template_path: str,
    output_path: str,
    year: int,
    month: int,
) -> None:
    """
    Write attendance records into the Excel template and save as a new file.

    Parameters
    ----------
    records : list[AttendanceRecord]
        Parsed attendance data (one per day).
    template_path : str
        Path to the Excel template (.xlsx).
    output_path : str
        Path where the filled Excel file will be saved.
    year : int
        The year of the attendance period.
    month : int
        The month of the attendance period.

    Raises
    ------
    FileNotFoundError
        If the template file does not exist.
    """
    if isinstance(template_path, str):
        if not os.path.isfile(template_path):
            raise FileNotFoundError(f"Template file not found: {template_path}")
        wb = load_workbook(template_path)
    else:
        wb = load_workbook(template_path)
    ws = wb.active

    # Configure print settings to E54 and 75% scale on portrait A4
    from openpyxl.worksheet.properties import PageSetupProperties
    ws.print_area = 'A1:E54'
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=False, autoPageBreaks=False)
    ws.page_setup.scale = 75
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToWidth = None
    ws.page_setup.fitToHeight = None

    # Ensure all images scale and move with cells (twoCell) so they resize correctly at 75% print scale
    for img in ws._images:
        if hasattr(img, 'anchor') and img.anchor:
            img.anchor.editAs = "twoCell"
            
            # Unmerge any merged cell ranges that contain the image's starting cell to prevent placement shifts in LibreOffice/HTML
            if type(img.anchor).__name__ == 'TwoCellAnchor' and img.anchor._from:
                start_row = img.anchor._from.row + 1
                start_col = img.anchor._from.col + 1
                for rng in list(ws.merged_cells.ranges):
                    if rng.min_row <= start_row <= rng.max_row and rng.min_col <= start_col <= rng.max_col:
                        log.info("Unmerging merged cell range %s overlapping with image anchor at row %d, col %d", 
                                 str(rng), start_row, start_col)
                        ws.unmerge_cells(str(rng))
                        break

    # Determine how many days this month has
    days_in_month = calendar.monthrange(year, month)[1]

    # Dynamically find the table position
    header_row, start_col = find_table_coordinates(ws)
    if header_row is None:
        log.info("Table header not found dynamically. Falling back to simple table layout (row 1, col 1).")
        start_row = 1
        start_col = 1
    else:
        start_row = header_row + 1
        log.info("Dynamic table placement detected at row %d, col %d.", start_row, start_col)
        # Also update the period cell if this is a full template
        update_periode_cell(ws, year, month)

    # Build a lookup dict: day number → record
    records_by_day: dict[int, AttendanceRecord] = {r.tgl: r for r in records}

    for day in range(1, days_in_month + 1):
        row = start_row + day - 1
        record = records_by_day.get(day)

        # Always unmerge first so we can write individual cells
        _unmerge_row(ws, row)

        # Write date number in start_col (with border)
        date_cell = ws.cell(row=row, column=start_col)
        date_cell.value = day
        date_cell.border = THIN_BORDER

        # --- Holiday / leave from HTML ---
        if record is not None and record.is_holiday_or_leave:
            log.debug("Day %d: holiday/leave → '%s'", day, record.keterangan)
            ws.cell(row=row, column=start_col + 1).value = record.keterangan
            for col in range(start_col + 2, start_col + 5):
                ws.cell(row=row, column=col).value = None
            _apply_holiday_style(ws, row, start_col)  # borders first, before merge
            _merge_be(ws, row, start_col)
            continue

        # --- Normal attendance data ---
        if record is not None and record.has_attendance:
            ws.cell(row=row, column=start_col + 1).value = record.masuk
            ws.cell(row=row, column=start_col + 2).value = record.pulang
            ws.cell(row=row, column=start_col + 3).value = None
            ws.cell(row=row, column=start_col + 4).value = None
            for col in range(start_col + 1, start_col + 5):
                _apply_normal_style(ws, row, col)
            log.debug("Day %d: wrote attendance data.", day)
            continue

        # --- No attendance and no remark: check for weekend ---
        weekend_label = _get_weekend_label(year, month, day)
        if weekend_label:
            log.debug("Day %d: weekend auto-detected → '%s'", day, weekend_label)
            ws.cell(row=row, column=start_col + 1).value = weekend_label
            for col in range(start_col + 2, start_col + 5):
                ws.cell(row=row, column=col).value = None
            _apply_holiday_style(ws, row, start_col)  # borders first, before merge
            _merge_be(ws, row, start_col)
            continue

        # --- Empty day (weekday, no attendance, no remark) ---
        log.debug("Day %d: no attendance, clearing cells.", day)
        for col in range(start_col + 1, start_col + 5):
            ws.cell(row=row, column=col).value = None
            _apply_normal_style(ws, row, col)

    # Clear any remaining rows beyond days_in_month (up to 31 days)
    for day in range(days_in_month + 1, 32):
        row = start_row + day - 1
        _unmerge_row(ws, row)
        tgl_cell = ws.cell(row=row, column=start_col)
        tgl_cell.value = None
        tgl_cell.border = THIN_BORDER
        for col in range(start_col + 1, start_col + 5):
            ws.cell(row=row, column=col).value = None
            _apply_normal_style(ws, row, col)

    # Save to output path
    wb.save(output_path)
    log.info("Output saved to: %s", output_path)


# ---------------------------------------------------------------------------
# CLI Entry Point
# ---------------------------------------------------------------------------
def main() -> None:
    """Main entry point with argument parsing."""
    parser = argparse.ArgumentParser(
        description="Process attendance HTML data into an Excel template.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""\
Example:
  python3 process_absen.py \\
      --html "Absen Bulan 03-2026.html" \\
      --template template-table-absen.xlsx \\
      --output output_absen.xlsx
""",
    )
    parser.add_argument(
        "--html",
        required=True,
        help="Path to the HTML attendance file.",
    )
    parser.add_argument(
        "--template",
        required=True,
        help="Path to the Excel template file (.xlsx).",
    )
    parser.add_argument(
        "--output",
        default="output_absen.xlsx",
        help="Output Excel file path (default: output_absen.xlsx).",
    )
    parser.add_argument(
        "--verbose", "-v",
        action="store_true",
        help="Enable debug-level logging.",
    )

    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    try:
        # Step 1: Parse the HTML file
        records, year, month = parse_html(args.html)
        log.info("Parsed %d attendance records for %04d-%02d.",
                 len(records), year, month)

        # Step 2: Print summary
        attendance_days = sum(1 for r in records if r.has_attendance)
        holiday_days = sum(1 for r in records if r.is_holiday_or_leave)
        empty_days = sum(
            1 for r in records
            if not r.has_attendance and not r.is_holiday_or_leave
        )
        log.info(
            "Summary: %d attendance | %d holiday/leave | %d empty",
            attendance_days, holiday_days, empty_days,
        )

        # Step 3: Write to Excel
        write_to_excel(records, args.template, args.output, year, month)

        print(f"\n✅ Done! Output saved to: {args.output}")

    except FileNotFoundError as e:
        log.error("File not found: %s", e)
        sys.exit(1)
    except ValueError as e:
        log.error("Parsing error: %s", e)
        sys.exit(2)
    except Exception as e:
        log.error("Unexpected error: %s", e, exc_info=True)
        sys.exit(3)


if __name__ == "__main__":
    main()
