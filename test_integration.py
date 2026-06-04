#!/usr/bin/env python3
"""
Integration Test Script
======================
Verifies the parsing and XLSX generation logic end-to-end without needing
to start the Flask HTTP server.
"""

import os
import sys
from openpyxl import load_workbook

# Ensure current directory is in path
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from process_absen import parse_html, write_to_excel, AttendanceRecord

def test_integration():
    print("=== Running Attendance Integration Test ===")
    
    html_input = "Absen Bulan 05-2026.html"
    simple_template = "template-table-absen.xlsx"
    full_template = "template_absen.xlsx"
    output_xlsx = "output_test_integration.xlsx"
    
    if not os.path.exists(html_input):
        print(f"❌ Error: HTML input file '{html_input}' not found in workspace.")
        return False
        
    # 1. Test HTML parsing (done once)
    print("\n1. Testing HTML parsing...")
    try:
        records, year, month = parse_html(html_input)
        print(f"   - Successfully parsed period: {year}-{month:02d}")
        print(f"   - Total records parsed: {len(records)}")
        
        # Verify first day
        day_1 = next((r for r in records if r.tgl == 1), None)
        print(f"   - Day 1 info: {day_1}")
        assert day_1 is not None, "Day 1 record is missing"
        assert day_1.is_holiday_or_leave, "Day 1 should be a holiday (Hari Buruh)"
        assert day_1.keterangan == "Hari Buruh Internasional", f"Expected 'Hari Buruh Internasional', got '{day_1.keterangan}'"
        
        # Verify day 5
        day_5 = next((r for r in records if r.tgl == 5), None)
        print(f"   - Day 5 info: {day_5}")
        assert day_5 is not None, "Day 5 record is missing"
        assert day_5.has_attendance, "Day 5 should have attendance"
        assert day_5.masuk == "07:58", f"Expected '07:58', got '{day_5.masuk}'"
        assert day_5.pulang == "17:11", f"Expected '17:11', got '{day_5.pulang}'"
        
        print("   ✅ HTML parsing tests passed!")
    except Exception as e:
        print(f"   ❌ HTML parsing failed: {e}")
        return False

    # 2. Test modification (Simulating user editing in Frontend - done once)
    print("\n2. Simulating frontend edits...")
    try:
        modified_records = []
        for r in records:
            masuk = r.masuk
            pulang = r.pulang
            keterangan = r.keterangan
            
            if r.tgl == 5:
                masuk = "08:15"
                pulang = "17:45"
                print(f"   - Edited Day 5 to: in={masuk} out={pulang}")
                
            if r.tgl == 8:
                masuk = None
                pulang = None
                keterangan = "CUTI BERSAMA (EDITED)"
                print(f"   - Edited Day 8 to: holiday={keterangan}")
                
            modified_records.append(AttendanceRecord(
                tgl=r.tgl,
                masuk=masuk,
                pulang=pulang,
                keterangan=keterangan
            ))
            
        print("   ✅ Simulation successful!")
    except Exception as e:
        print(f"   ❌ Edit simulation failed: {e}")
        return False

    # 3. Determine templates to test
    templates_to_test = []
    if os.path.exists(simple_template):
        templates_to_test.append(("Simple Table Template", simple_template))
    else:
        print(f"⚠️ Warning: Simple template '{simple_template}' not found.")
        
    if os.path.exists(full_template):
        templates_to_test.append(("Full Attendance Template", full_template))
    else:
        print(f"ℹ️ Info: '{full_template}' not found yet. Please make sure template_absen.xlsx is present in the workspace.")
        
    if not templates_to_test:
        print("❌ Error: No template files found to run integration tests.")
        return False

    # 4. Run template generation and verification for each template
    from process_absen import find_table_coordinates
    for name, template_xlsx in templates_to_test:
        print(f"\n--- Testing Template: {name} ({template_xlsx}) ---")
        try:
            if os.path.exists(output_xlsx):
                os.remove(output_xlsx)
                
            write_to_excel(
                records=modified_records,
                template_path=template_xlsx,
                output_path=output_xlsx,
                year=year,
                month=month
            )
            print(f"   - Successfully saved output to: {output_xlsx}")
            assert os.path.exists(output_xlsx), "Output file was not created"
            
            # Open generated workbook and inspect values
            wb = load_workbook(output_xlsx)
            ws = wb.active
            
            # Find table offsets dynamically
            h_row, s_col = find_table_coordinates(ws)
            if h_row is None:
                day_5_row = 5
                day_5_masuk_col = 2
                day_5_pulang_col = 3
                day_8_row = 8
                day_8_masuk_col = 2
            else:
                day_5_row = h_row + 5
                day_5_masuk_col = s_col + 1
                day_5_pulang_col = s_col + 2
                day_8_row = h_row + 8
                day_8_masuk_col = s_col + 1
                
            # Verify Day 5 edited times
            val_masuk_5 = ws.cell(row=day_5_row, column=day_5_masuk_col).value
            val_pulang_5 = ws.cell(row=day_5_row, column=day_5_pulang_col).value
            print(f"   - Excel Day 5: Row {day_5_row}, Col {day_5_masuk_col}='{val_masuk_5}' Col {day_5_pulang_col}='{val_pulang_5}'")
            assert val_masuk_5 == "08:15", f"Expected '08:15', got '{val_masuk_5}'"
            assert val_pulang_5 == "17:45", f"Expected '17:45', got '{val_pulang_5}'"
            
            # Verify Day 8 edited holiday remark
            val_masuk_8 = ws.cell(row=day_8_row, column=day_8_masuk_col).value
            print(f"   - Excel Day 8: Row {day_8_row}, Col {day_8_masuk_col}='{val_masuk_8}'")
            assert val_masuk_8 == "CUTI BERSAMA (EDITED)", f"Expected 'CUTI BERSAMA (EDITED)', got '{val_masuk_8}'"
            
            # If it is the full template, verify the period cell got updated
            if h_row is not None:
                # We expect the word "JUNI" or "MEI" in the period cell depending on month
                # Let's search for the updated period cell in rows 1 to h_row
                periode_found = False
                for r in range(1, h_row):
                    for c in range(1, 15):
                        cell_val = ws.cell(row=r, column=c).value
                        if cell_val and "periode" in str(cell_val).lower():
                            print(f"   - Excel Periode label cell: Row {r}, Col {c}='{cell_val}'")
                            # Verify if period is in the cell itself
                            if "2026" in str(cell_val) and "1 -" in str(cell_val):
                                periode_found = True
                                break
                            # Or check adjacent cells to the right
                            for next_c in range(c + 1, c + 7):
                                right_val = ws.cell(row=r, column=next_c).value
                                if right_val and "2026" in str(right_val) and "1 -" in str(right_val):
                                    print(f"   - Excel Periode value cell: Row {r}, Col {next_c}='{right_val}'")
                                    periode_found = True
                                    break
                        if periode_found:
                            break
                    if periode_found:
                        break
                assert periode_found, "Could not locate updated period cell in full template output"
                
            wb.close()
            
            # Clean up test output
            if os.path.exists(output_xlsx):
                os.remove(output_xlsx)
                print("   - Cleaned up output test file.")
                
            print(f"   ✅ tests passed for {name}!")
        except Exception as e:
            print(f"   ❌ tests failed for {name}: {e}")
            import traceback
            traceback.print_exc()
            return False
            
    print("\n🎉 ALL INTEGRATION TESTS COMPLETED SUCCESSFULLY!")
    return True

def test_endpoints():
    print("\n=== Running Flask API Endpoint Mock Tests ===")
    from unittest.mock import patch, MagicMock
    from app import app
    
    client = app.test_client()
    
    # 1. Test Health endpoint
    print("1. Testing /api/health endpoint...")
    res = client.get('/api/health')
    assert res.status_code == 200
    health_data = res.get_json()
    assert health_data["status"] == "healthy"
    print("   ✅ Health endpoint test passed!")

    # 2. Test Login endpoint (mocked)
    print("2. Testing /api/login endpoint with mocks...")
    with patch('app.requests.Session') as mock_session_cls:
        mock_session = MagicMock()
        mock_session_cls.return_value = mock_session
        
        # Mock GET login page
        mock_get_login = MagicMock()
        mock_get_login.status_code = 200
        mock_get_login.text = '<html><meta name="csrf-token" content="mock-csrf-token-123"></html>'
        
        # Mock POST login submission (returns 302 Redirect)
        mock_post_login = MagicMock()
        mock_post_login.status_code = 302
        mock_post_login.text = 'Redirecting...'
        
        # Mock GET profile home page
        mock_get_home = MagicMock()
        mock_get_home.status_code = 200
        mock_get_home.text = '<html><body><button class="logout">Logout (WIBI CHAMIM MUSHODIQ)</button></body></html>'
        
        # Define session call side effects
        mock_session.get.side_effect = [mock_get_login, mock_get_home]
        mock_session.post.return_value = mock_post_login
        
        # Mock session cookies
        mock_session.cookies.get_dict.return_value = {
            "PHPSESSID": "mock-php-sessid",
            "_identity-absen-bisnis": "mock-identity-token",
            "_csrf-absen-bisnis": "mock-csrf-token-123"
        }
        
        # Request login
        res = client.post('/api/login', json={
            "username": "3259800588",
            "password": "secretpassword"
        })
        
        assert res.status_code == 200, f"Expected 200, got {res.status_code}"
        data = res.get_json()
        
        print(f"   - Response payload: {data}")
        assert data["status"] == "success"
        assert data["display_name"] == "WIBI CHAMIM MUSHODIQ"
        assert "PHPSESSID=mock-php-sessid" in data["session_cookie"]
        assert "_identity-absen-bisnis=mock-identity-token" in data["session_cookie"]
        
        # Assert parameters passed to post
        called_args, called_kwargs = mock_session.post.call_args
        assert called_kwargs["data"]["LoginForm[username]"] == "3259800588"
        assert called_kwargs["data"]["_csrf-absen-bisnis"] == "mock-csrf-token-123"
        
        print("   ✅ Mocked login endpoint tests passed!")

        # 3. Test Template Endpoints
        print("\n3. Testing template endpoints...")
        # Get status
        res = client.get('/api/template-status')
        assert res.status_code == 200
        status_data = res.get_json()
        assert "active_template" in status_data
        assert "is_custom" in status_data
        print(f"   - Initial active template: {status_data['active_template']}")

        # Download template
        res = client.get('/api/download-template')
        assert res.status_code in (200, 404)
        if res.status_code == 200:
            assert res.headers["Content-Disposition"].startswith("attachment")
            print("   - Template download successful")

        # Upload custom template
        import io
        import openpyxl
        wb = openpyxl.Workbook()
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        res = client.post('/api/upload-template', data={
            'file': (excel_file, 'test_custom_template.xlsx')
        }, content_type='multipart/form-data')
        assert res.status_code == 200
        upload_data = res.get_json()
        assert upload_data["status"] == "success"
        assert upload_data["active_template"] == "template_custom.xlsx"
        print("   - Upload custom template successful")

        # Re-verify template status
        res = client.get('/api/template-status')
        assert res.status_code == 200
        status_data = res.get_json()
        assert status_data["is_custom"] is True
        assert status_data["active_template"] == "template_custom.xlsx"

        # Reset template
        res = client.post('/api/reset-template')
        assert res.status_code == 200
        reset_data = res.get_json()
        assert reset_data["status"] == "success"
        print("   - Reset custom template successful")

        # Re-verify template status is default
        res = client.get('/api/template-status')
        assert res.status_code == 200
        status_data = res.get_json()
        assert status_data["is_custom"] is False
        print("   ✅ Template endpoints tests passed!")

        # 4. Test Preview Endpoint
        print("\n4. Testing /api/preview endpoint...")
        # Since xlsx2html might not be installed in the sandboxed run, we handle both outcomes
        res = client.post('/api/preview', json={
            "year": 2026,
            "month": 5,
            "records": [
                {"tgl": 1, "masuk": "08:00", "pulang": "17:00", "keterangan": None}
            ]
        })
        assert res.status_code in (200, 500)
        if res.status_code == 200:
            preview_data = res.get_json()
            assert "html" in preview_data
            print("   - Preview HTML rendering successful")
        else:
            err_data = res.get_json()
            assert "xlsx2html" in err_data.get("error", "")
            print("   - Preview returned expected dependency missing error (fine for sandbox)")

        # 5. Test Generate PDF Endpoint
        print("\n5. Testing /api/generate-pdf endpoint...")
        res = client.post('/api/generate-pdf', json={
            "year": 2026,
            "month": 5,
            "records": [
                {"tgl": 1, "masuk": "08:00", "pulang": "17:00", "keterangan": None}
            ]
        })
        assert res.status_code in (200, 400)
        if res.status_code == 200:
            assert res.mimetype == "application/pdf"
            print("   - PDF download successful")
        else:
            pdf_data = res.get_json()
            assert "LibreOffice" in pdf_data["error"]
            print("   - Direct PDF returned expected LibreOffice missing error (fine for sandbox)")

        # 6. Test Remote Attendance Endpoint
        print("\n6. Testing /api/remote-absen endpoint...")
        mock_get_remote = MagicMock()
        mock_get_remote.status_code = 200
        mock_get_remote.text = '<html><meta name="csrf-token" content="mock-csrf-token-abc"><body><button class="logout">Logout (WIBI CHAMIM MUSHODIQ)</button></body></html>'
        
        mock_post_remote = MagicMock()
        mock_post_remote.status_code = 200
        mock_post_remote.text = '<html><body><div class="alert-success">Absensi remote berhasil dilakukan.</div></body></html>'
        mock_post_remote.url = 'https://ksps.co.id/eksternal/absen/index'
        
        mock_session.get.side_effect = [mock_get_remote]
        mock_session.post.return_value = mock_post_remote
        
        res = client.post('/api/remote-absen', json={
            "session_cookie": "PHPSESSID=mock-php-sessid; _identity-absen-bisnis=mock-identity-token",
            "latitude": "-7.7837217165",
            "longitude": "110.4329516476",
            "status": "0",
            "device_token": "mock-device-token-xyz"
        })
        assert res.status_code == 200
        remote_data = res.get_json()
        assert remote_data["status"] == "success"
        assert "berhasil" in remote_data["message"]
        
        # Assert parameters passed to post for remote absen
        called_args, called_kwargs = mock_session.post.call_args
        assert called_kwargs["data"]["Absen[token]"] == "mock-device-token-xyz"
        assert called_kwargs["data"]["Absen[lintang]"] == "-7.7837217165"
        
        print("   - Remote check-in mock test successful")



        return True

if __name__ == "__main__":
    success = test_integration() and test_endpoints()
    sys.exit(0 if success else 1)
